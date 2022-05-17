import keepa
import json
import openpyxl.drawing.image
import numpy as np
import PIL.Image
import openpyxl
import time
import datetime
import urllib.request
from openpyxl.styles import Alignment

# APIキーを読み込み
file=open('config.json',"r")
j_file=json.load(file)
accesskey=j_file["API_KEY"]
api=keepa.Keepa(accesskey)

# 出力用日付データを取得
dt_now=datetime.datetime.now()
dtnf=dt_now.strftime('%Y%m%d_%H%M%S')

global products
global dst_path
title=""
asins=[]

# Excel読込
wb=openpyxl.load_workbook("asin_to_xl.xlsx") 
ws=wb['general']
url="https://m.media-amazon.com/images/I/"
print("ASINを抽出します。Excelは必ず閉じてください")
time.sleep(1)


def get_asins():
    print("ASINの読取中")
# 変な値が入っていたら値をnoneにする。
    for row in ws.iter_rows(min_row=ws.max_row+1, min_col=1, max_row=5000, max_col=1):
        for cell in row:
            cell.value = None
# ExcelからASINを読み取り、リストasinsに格納
    for row in range(2,ws.max_row+1):
        if ws.cell(row,column=1).value==None:
            pass
        else:
            asin=ws.cell(row,column=1).value
            asins.append(asin.replace(" ",""))
    print("ASINの読取完了")

'''
画像データの読込用関数　本案件では使用しない
def download_file(name):
    dst_path="./{}.png".format(dtnf)
    url=f"https://m.media-amazon.com/images/I/{name}"
    with urllib.request.urlopen(url) as web_file:
        data=web_file.read()
        with open(dst_path,mode="wb") as local_file:
            local_file.write(data) 
def resize_img(dst_path):
    datam=PIL.Image.open(dst_path)
    data_resize=datam.resize((90,90))
    data_resize.save(dst_path)
def paste_toxslx(dst_path,position):
    img_to_excel=openpyxl.drawing.image.Image(dst_path)
    ws.add_image(img_to_excel,position)
'''

def to_excel():
    print("商品情報を呼び出します")
    print(f"商品数量は{len(asins)}件です。")
    
    # keepa apiを叩く。productsにオブジェクト格納。
    products=api.query(asins,domain='JP')
    for i,product in enumerate(products,2):
        
        '''
        商品画像の読み込み用。本案件では使用しない
        if product['imagesCSV']==None:
            print("なし")
        else:
            img_list=product['imagesCSV'].split(",")
            name=img_list[0]
            dst_path="./{}.png".format(dtnf)
            url=f"https://m.media-amazon.com/images/I/{name}"
            with urllib.request.urlopen(url) as web_file:
                data=web_file.read()
                with open(dst_path,mode="wb") as local_file:
                    local_file.write(data) 
            resize_img(dst_path) 
            paste_toxslx(dst_path,f'b{i}')
            wb.save("asin_to_xl.xlsx")
        '''
        
        pro_asin=product['asin']
        for r in range(2,ws.max_row+1):
            if pro_asin==ws.cell(r,1).value:
                # JANコード
                try:
                    if product['eanList']==None:
                        ws.cell(r,2).value="なし"
                    else:
                        
                        ean_l=[]
                        try:
                            for k in range(0,len(product['eanList'])):
                                ean_l.append(product['eanList'][k])
                            for l in range(0,len(product['upcList'])):
                                ean_l.append(product['eanList'][l])
                        except:
                            pass
                        ean_lj="\n".join(ean_l)
                        ws.cell(r,2).value=ean_lj
                        ws.cell(r,2).number_format="0"
                        ws.cell(r,2).alignment=Alignment(horizontal="left",vertical="center")
                        
                        # 選択範囲内で中央を設定したい
                except:
                    pass    
            
                # 商品タイトル
                try:
                    if product['title']==None:
                        ws.cell(r,3).value="なし"
                    else:
                        ws.cell(r,3).value=product['title']
                except:
                    pass
                
                # この商品について
                try:
                    if product['description']==None:
                        ws.cell(r,4).value="なし"
                    else:
                        pro_des=product['description']
                        ws.cell(r,4).value=pro_des.replace("     ","\n")
                        # ws.cell(i,4).value="\n".join(pro_des)
                except:
                    pass
                
                # 商品の説明
                if product['features']==None:
                    ws.cell(r,5).value="なし"
                else:
                    ws.cell(r,5).value="\n".join(product['features'])
            else:
                pass

    print("Excelへの書込完了。ツールを終了します。")
    ws.column_dimensions['B'].width=20
    wb.save(f"出力結果{dtnf}.xlsx")
    
get_asins()
to_excel()
print(f"残りのトークン数は{api.tokens_left}です。")
time.sleep(2)
